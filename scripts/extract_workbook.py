#!/usr/bin/env python3
"""Converte a planilha oficial do simulador nos arquivos JS consumidos pelo site."""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def number(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, "", "sob consulta"):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def write_js(path: Path, variable: str, payload: dict[str, Any]) -> None:
    path.write_text(
        f"{variable}=" + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )


def build(workbook_path: Path, output_dir: Path) -> None:
    wb = load_workbook(workbook_path, data_only=True)

    # Clientes
    clientes = []
    ws = wb["clientes"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:
            continue
        clientes.append({
            "sigla": text(row[0]), "instituicao": text(row[1]), "esfera": text(row[2]),
            "cnpj": text(row[3]), "representante": text(row[4]), "cargo": text(row[5]),
            "telefone": text(row[6]), "endereco": text(row[7]), "bairro": text(row[8]),
            "cep": text(row[9]), "municipio": text(row[10]), "uf": text(row[11]),
        })

    # Serviços
    servicos = []
    ws = wb["servicos"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[3]:
            continue
        anual_raw = row[12]
        mensal_raw = row[14]
        sob_consulta = any(text(v).lower() == "sob consulta" for v in (row[9], row[11], row[12]))
        servicos.append({
            "area": text(row[0]), "subarea": text(row[1]), "codigo": text(clean(row[2])),
            "descricao": text(row[3]), "tipo": text(row[4]), "grandeza": text(row[8]),
            "faturamento": text(row[10]), "precoMensal": number(mensal_raw),
            "precoAnual": number(anual_raw), "sobConsulta": sob_consulta,
        })

    # Regiões: a tabela principal começa na linha 6 e tem 145 municípios.
    regioes = []
    ws = wb["regiao"]
    for row in ws.iter_rows(min_row=6, values_only=True):
        municipio = text(row[0])
        if not municipio:
            continue
        # Evita capturar tabelas auxiliares à direita/abaixo da tabela municipal.
        if not text(row[3]) and not text(row[5]):
            continue
        regioes.append({
            "municipio": municipio, "backbone": text(row[1]), "ultimaMilha": text(row[2]),
            "atendido": text(row[3]), "mesorregiao": text(row[4]), "regiao": text(row[5]),
            "atualizacao": text(row[6]), "precoMbps": number(row[7]),
        })

    # Ajustes emergenciais sem regravar o XLSX (preserva fórmulas e vínculos da planilha).
    overrides_path = workbook_path.with_name("overrides.json")
    if overrides_path.exists():
        overrides = json.loads(overrides_path.read_text(encoding="utf-8"))
        region_overrides = overrides.get("regioes", {})
        for item in regioes:
            patch = region_overrides.get(item["municipio"])
            if patch:
                item.update(patch)

    # Configurações e descontos
    ws = wb["pconfig"]
    reajustes = {}
    for r in range(2, 7):
        ano = ws.cell(r, 1).value
        if ano:
            reajustes[str(int(ano))] = number(ws.cell(r, 2).value)

    faixas = []
    for r in range(12, ws.max_row + 1):
        minimo, maximo, fator = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if minimo is None or maximo is None or fator is None:
            if faixas:
                break
            continue
        faixas.append({"min": number(minimo), "max": number(maximo), "fator": number(fator)})

    config = {
        "taxaAdministrativa": number(ws["G2"].value),
        "reaparelhamento": number(ws["G3"].value),
        "impostos": number(ws["G4"].value),
        "ativacao": number(ws["G7"].value),
        "configFibra": number(ws["G8"].value),
        "configRadio": number(ws["G9"].value),
        "internetMbps": number(ws["G14"].value, 23.0),
    }

    # Link de Dados
    link_dados = []
    ws = wb["linkdados"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[2]:
            continue
        link_dados.append({
            "municipio": text(row[1]), "modalidade": text(row[2]), "tecnologias": text(row[3]),
            "mesorregiao": text(row[6]), "atendido": text(row[7]), "regiao": text(row[8]),
            "degrau": text(row[9]), "unitTransporte2026": number(row[11]),
            "cmanut2026": number(row[13]), "precoFinal1Mbps2026": number(row[14]),
        })

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "dados").mkdir(parents=True, exist_ok=True)
    write_js(output_dir / "dados" / "data.js", "window.APP_DATA", {
        "clientes": clientes, "servicos": servicos, "regioes": regioes, "config": config,
    })
    write_js(output_dir / "dados" / "pricing.js", "window.LINK_DATA_PRICING", {
        "reajustes": reajustes, "faixasDesconto": faixas, "linkDados": link_dados,
    })
    print(f"Gerados data.js e pricing.js a partir de {workbook_path}")
    print(f"Clientes: {len(clientes)} | Serviços: {len(servicos)} | Municípios: {len(regioes)} | Preços: {len(link_dados)}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default="planilhas/simulador-base.xlsx")
    parser.add_argument("--output", default=".")
    args = parser.parse_args()
    build(Path(args.workbook), Path(args.output))


if __name__ == "__main__":
    main()
